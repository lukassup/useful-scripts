#!/usr/bin/env python
# -*- coding: utf-8 -*-

from csv import reader
from openpyxl import Workbook
import os.path
from sys import (
        exit,
        argv,
        stderr,
        stdout,
        stdin
        )

DEBUG = True

def debug(*mesg):
    if DEBUG:
        print("***", *mesg, file=stderr)


def build_worksheet(worksheet, csv_file):
    csv_reader = reader(csv_file)
    for row in csv_reader:
        for x, item in enumerate(row, start=1):
            worksheet.cell(
                    row = csv_reader.line_num,
                    column = x,
                    value = item)

    debug("{0} rows written to sheet '{1}'"
            .format(csv_reader.line_num, worksheet.title))


def main():
    if not len(argv[1:]) > 0:
        debug("No files provided")
        return 1

    for arg in argv[1:]:
        if not os.path.exists(arg):
            debug("Path '{0}' does not exist or no permission.".format(arg))
            return 1

    wb = Workbook()
    debug("Workbook created")
    wb.remove_sheet(wb.active)

    for current_file in argv[1:]:
        debug("Working on '{0}'".format(current_file))

        sheet_name = os.path.splitext(
                        os.path.basename(
                            current_file))[0]

        debug("Adding sheet '{0}'".format(sheet_name))
        ws = wb.create_sheet(title=sheet_name)

        with open(current_file, "r") as data_file:
            debug("Building worksheet '{0}'".format(ws.title))
            build_worksheet(ws, data_file)

    wb.save(input(">>> Merged file name without \".xlsx\": ") + ".xlsx")
    debug("File saved")
    return 0


if __name__ == "__main__":
    try:
        exit(main())
    except EOFError:
        print(file=stderr)
        debug("Aborted")
    except KeyboardInterrupt:
        print(file=stderr)
        debug("Aborted")
    except Exception as e:
        debug(e)
    finally:
        exit(1)
